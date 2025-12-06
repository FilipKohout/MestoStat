import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from typing import Any, Iterator, Dict


def _iter_rows(sheet: Worksheet) -> Iterator[Dict[str, Any]]:
    headers = [cell.value for cell in sheet[1]]

    unique_headers = []
    header_counts = {}
    for header in headers:
        if header is None:
            base_header = f"Column_{len(unique_headers) + 1}"
        else:
            base_header = str(header)

        if base_header in header_counts:
            header_counts[base_header] += 1
            unique_headers.append(f"{base_header}_{header_counts[base_header]}")
        else:
            header_counts[base_header] = 0
            unique_headers.append(base_header)

    for row_idx in range(2, sheet.max_row + 1):  # Start from the second row for data
        row_data: Dict[str, Any] = {}
        for col_idx, header in enumerate(unique_headers):
            cell_value = sheet.cell(row=row_idx, column=col_idx + 1).value
            row_data[header] = cell_value
        yield row_data


def decode(file_path: str) -> dict[str, list[dict]]:
    """
    Decodes an XLSX file into a dictionary where keys are sheet names
    and values are lists of dictionaries (rows).
    """
    results: dict[str, list[dict]] = {}
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)  # data_only=True to get cell values, not formulas
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            rows = list(_iter_rows(sheet))
            results[sheet_name] = rows
    except Exception as e:
        # Handle cases where the file might be corrupted or not a valid XLSX
        print(f"Error decoding XLSX file {file_path}: {e}")
        # Optionally, you might want to log this error and return an empty dict or re-raise
        return {}  # Return empty if decoding fails

    return results